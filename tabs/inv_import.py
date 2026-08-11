#!/usr/bin/env python3
"""Reads an xlsx inventory file and outputs JSON rows per sheet (A1-A8)."""
import sys, json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

SHEETS = ['A1','A2','A3','A4','A5','A6','A7','A8']

def cell_str(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()

def excel_date(val):
    """Convert Excel serial date to YYYY-MM-DD, or return None."""
    try:
        n = float(val)
        if n > 40000:
            return datetime.fromordinal(datetime(1900,1,1).toordinal() + int(n) - 2).strftime('%Y-%m-%d')
    except Exception:
        pass
    return None

filepath = sys.argv[1] if len(sys.argv) > 1 else ''
if not filepath:
    print(json.dumps({"error": "no filepath given"}))
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(filepath, data_only=True)
except Exception as e:
    print(json.dumps({"error": str(e)}))
    sys.exit(1)

result = {}
for arm in SHEETS:
    if arm not in wb.sheetnames:
        continue
    ws = wb[arm]
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        desc = cell_str(row[1] if len(row) > 1 else None)
        if not desc:
            continue
        qty  = cell_str(row[2] if len(row) > 2 else None)
        prat = cell_str(row[3] if len(row) > 3 else None)
        cx   = cell_str(row[4] if len(row) > 4 else None)
        proj = cell_str(row[5] if len(row) > 5 else None)

        raw6 = row[6] if len(row) > 6 else None
        if isinstance(raw6, datetime):
            led = raw6.strftime('%Y-%m-%d')
        elif raw6 is not None and str(raw6).strip():
            led = excel_date(raw6)
        else:
            led = None

        rows.append({'descricao':desc,'quantidade':qty,'prateleira':prat,'caixa':cx,'projeto':proj,'last_edited':led})
    result[arm] = rows

wb.close()
print(json.dumps(result))
